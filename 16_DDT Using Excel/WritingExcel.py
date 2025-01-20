import os

import openpyxl
#File-->Workbook--->sheet-->rows---cells

file= r"C:\Users\ruchika\PycharmProjects\Selenium_09_01\16_DDT Using Excel\Test.xlsx"
# file=os.getcwd() + r"\Test.xlsx"

workbook=openpyxl.load_workbook(file)
sheet =workbook.active    #or  sheet=workbook["Data"] or sheet=workbook["Sheet1"]

#this will create  5 rows and 3 cols but with same data
# 5 rows and 3 cols
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value='Welcome'

workbook.save(file) # in case of writing the file.


# Create the file with multiple data

# sheet.cell(1,1).value=123
# sheet.cell(1,2).value="Scott"
# sheet.cell(1,3).value="USA"
#
# sheet.cell(2,1).value=541
# sheet.cell(2,2).value="David"
# sheet.cell(2,3).value="Canada"
#
# sheet.cell(3,1).value=541
# sheet.cell(3,2).value="Mary"
# sheet.cell(3,3).value="New york"
#
# workbook.save(file)



