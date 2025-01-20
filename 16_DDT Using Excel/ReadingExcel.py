import os

import  openpyxl #package is required to perform DDT

#File-->Workbook--->sheet-->rows---cells

#file= r"C:\Users\ruchika\PycharmProjects\Selenium_09_01\16_DDT Using Excel\test.xlsx"
file=os.getcwd() + r"\data.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row # count no of rows in excel
print("The no of rows are", rows) #6
cols=sheet.max_column # count no of cols in excel
print("The no of cols are",cols)

#Reading all the rows and col from excel

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="     ")
    print()


