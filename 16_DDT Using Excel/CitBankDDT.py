import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)

driver = webdriver.Chrome()
driver.get("https://www.cit.com/cit-bank/resources/calculators/certificate-of-deposit-calculator")
driver.maximize_window()
# C:\Users\ruchika\PycharmProjects\Selenium_09_01\16_DDT Using Excel\CitData.xlsx
# C:\Users\ruchika\PycharmProjects\Selenium_09_01\16_DDT Using Excel\CitData.xlsx

file = os.getcwd() + r"\Cal1_Data.xlsx"
print(file)
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row
driver.find_element(By.ID, "mat-input-0").clear()
driver.find_element(By.ID, "mat-input-1").clear()
driver.find_element(By.ID, "mat-input-2").clear()
for r in range(2, rows + 1):
    deposit = readData(file, "Sheet1", r, 1)
    length = readData(file, "Sheet1", r, 2)
    rate = readData(file, "Sheet1", r, 3)
    compounding = readData(file, "Sheet1", r, 4)
    exp_mvalue = readData(file, "Sheet1", r, 5)


    driver.find_element(By.ID, "mat-input-0").send_keys(deposit)


    driver.find_element(By.ID, "mat-input-1").send_keys(length)


    driver.find_element(By.ID, "mat-input-2").send_keys(rate)

    compounding_dropdown = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//mat-form-field//mat-select"))
    )
    compounding_dropdown.click()
    time.sleep(2)

    compounding_options = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//mat-option//span"))
    )

    for option in compounding_options:
        if option.text.strip() == compounding:
            option.click()
            break
    time.sleep(2)

    calc_button =driver.find_element(By.XPATH, '//*[@id="CIT-chart-submit"]')
    # calc_button = driver.execute_script(
    #     "return document.evaluate(arguments[0], document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;",
    #     calc_button_xpath
    # )
    driver.execute_script("arguments[0].click();", calc_button)
    time.sleep(2)

    # actual = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.XPATH, "//span[@class='cd-calculator__maturity-value-number']"))
    # )

    time.sleep(5)
    actual_mvalue= driver.find_element(By.XPATH,'//*[@id="displayTotalValue"]').text
    actual_mvalue = actual_mvalue.replace("$", "").replace(",", "")
    print(exp_mvalue)
    print(actual_mvalue)
    if float(exp_mvalue) == float(actual_mvalue):
        result = "Passed"
        writeData(file, "Sheet1", r, 7, result)
    else:
        result = "Failed"
        writeData(file, "Sheet1", r, 7, result)


    driver.find_element(By.ID, "mat-input-0").clear()
    driver.find_element(By.ID, "mat-input-1").clear()
    driver.find_element(By.ID, "mat-input-2").clear()


driver.quit()